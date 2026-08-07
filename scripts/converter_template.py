import os
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLS = os.path.join(RAIZ, "白紙 FUJIARTE Ficha Cadastral Jun2024 (1).xls")
OUT_DIR = os.path.join(RAIZ, "app", "public", "templates")
OUT_XLSX = os.path.join(OUT_DIR, "ficha_fujiarte_template.xlsx")

def converter():
    print("Lendo template XLS...")
    rb = xlrd.open_workbook(SRC_XLS, formatting_info=True)
    rsheet = rb.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FICHA CADASTRAL"
    ws.views.sheetView[0].showGridLines = True

    # Copiar dimensoes e valores
    for r in range(rsheet.nrows):
        for c in range(rsheet.ncols):
            val = rsheet.cell_value(r, c)
            cell = ws.cell(row=r+1, column=c+1)
            cell.value = val

    # Copiar mesclagens de celulas
    for crange in rsheet.merged_cells:
        rlow, rhigh, clow, chigh = crange
        ws.merge_cells(
            start_row=rlow+1,
            start_column=clow+1,
            end_row=rhigh,
            end_column=chigh
        )

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Template XLSX salvo em: {OUT_XLSX}")

if __name__ == "__main__":
    converter()
